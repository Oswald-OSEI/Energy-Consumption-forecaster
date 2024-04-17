
# @AUTHOR = OSEI OSWALD OWUSU-ASIEDU

# TOPIC = PREDICTIVE ANYALYTICS AND FORECASTING OF ENERGY CONSUMPTION
import numpy as np
import openpyxl
from openpyxl import load_workbook
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn import linear_model

location = "input excel file location here" #stores the location of the excel file to be read

#INITIALIZING PARAMETERS AND VARIABLES
model_obj = linear_model.LinearRegression()
consumptions = []
days = []
accumulated_days = 0
linear_consumptions = []
linear_days = []
update = []
Per_error_list = []
accumulated_consumptions = 0
current_consumption = 0
update_consumption = 0

#READING THE EXCEL FILE
Work_Book = openpyxl.load_workbook(location)
sheet_current = Work_Book.active
n_column = sheet_current.max_column
n_row = sheet_current.max_row

# ACCUMULATING THE CONSUMPTIONS
for i in range(1, n_column + 1, 1):
    n_row = 1
    while n_row <= 12:
        cell_current = sheet_current.cell(row=n_row, column=i)
        monthly = cell_current.value
        n_row += 1
        accumulated_days += 30
        days.append(accumulated_days)
        consumptions.append(monthly)
for j in range(0, 1000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000, 1):
        cummulative = []
        for i in consumptions:
            accumulated_consumptions += i
            cummulative.append(accumulated_consumptions)
        print(cummulative)
        print(days)
        if j >= 1:
            # THIS SECTION ALLOWS FOR UPDATES OF THE CONTINUOUS CONSUMPTION AND DAYS OR MONTHS.
            # EVERY CONSUMPTION UPDATE IS DONE ON A MONTHLY BASIS WHERE EVERY MONTH IS ASSUMED TO HAVE 30DAYS.
            recent_consumption = float(input('enter your recent consumption:'))
            last_id = int((len(days))) - 1
            updated = 30 + float(days[last_id])
            days.append(updated)
            #print(days[last_id])
            print(last_id)
            update_consumption = recent_consumption + cummulative[last_id]
            cummulative.append(update_consumption)
            print(cummulative)
            print(cummulative[last_id +1])

        estimate = int((len(days))) - 1
        #converting the list to a 1 dimensional array usngin numpy
        x = np.array(days)
        y = np.array(cummulative)
        #fitting the degree of polynomial and giving it as well as the linear regressor function a variable
        poly = PolynomialFeatures(degree=3)
        pilreg = LinearRegression()
        #convert the X array from 1 dimension to 2 dimensional array
        X = poly.fit_transform(x.reshape(-1, 1))
        #splitting data into training and testing data set
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size= 0.2, random_state=1)
        print('Your training dataset for independent data is', X_train)
        print('Your training dataset for dependent data is', y_train)
        print('Your testing dataset for independent data is', X_test)
        print('Your testing dataset for dependent data is', y_test)

        #fitting the train data set to the model
        pilreg.fit(X_train, y_train)
        actual_con = (y_test).tolist()
        test_pred = (pilreg.predict(X_test)).tolist()
        print('Your predicted output on the various testing independent values are', pilreg.predict(X_test))
        #Model evaluation
        #calculating the percentage error and writing unto excel file
        #calculating error
        v = len(test_pred)
        for k in range(0, v, 1):
            Per_error = ((abs(float(test_pred[k]) -float(actual_con[k])))/actual_con[k])*100
            Per_error_list.append(Per_error)
        print('Percentage error = ', Per_error_list)
        #writing unto excel file
        wb = load_workbook('evaluation.xlsx')
        active_sheet = wb.create_sheet(location, -1)
        active_sheet.title = location
        #Setting column width
        active_sheet.column_dimensions['A'].width = 25
        active_sheet.column_dimensions['B'].width = 25
        active_sheet.column_dimensions['C'].width = 25
        (active_sheet.cell(row= 1, column = 1)).value = "Actual Consumption"
        (active_sheet.cell(row=1, column=2)).value = "Predicted Consumption"
        (active_sheet.cell(row=1, column=3)).value = "% Error"
        b=2
        for s in range (0, v, 1):
            (active_sheet.cell(row=b, column=1, )).value = actual_con[s]
            (active_sheet.cell(row=b, column=2)).value = test_pred[s]
            (active_sheet.cell(row=b, column=3)).value = Per_error_list[s]
            b +=1

        wb.save('evaluation.xlsx')



        #Calculating the model accuracy
        print(' models coefficient of determination  for ', (location), 'is', pilreg.score(X_train, y_train))
        print(' model accuracy on unseen data(test dataset) for', (location), 'is', pilreg.score(X_test, y_test))
        #Asking user to input the number of days he would want to budget for
        month = input('enter the number of days: ')
        check_days = int(month) + x[estimate]
        months = int(month)
        print(check_days)
        prediction1 = pilreg.predict(poly.fit_transform([[check_days]]))
        previous = pilreg.predict(poly.fit_transform([[x[estimate]]]))
        actual = pilreg.predict(poly.fit_transform([[months]]))
        #print(prediction1)
        #print(previous)
        prediction= abs(prediction1 - previous)
        #print(prediction)
        #outputing estimated usage per days input
        print('ESTIMATED CONSUMPTION TO BE BUDGETED FOR',(location),  'SHOULD BE AROUND',prediction,'kwh')
        #R = r2_score(y_train, pilreg.predict(X_train))
        #print(R)


        #THE FOLLOWING CODES ESTIMATES HOW LONG YOUR REMAINING YOUR CREDIT WILL LAST
        CREDIT = float(input('Enter your remaining meter credit:'))
        monthly = int(30 + x[estimate])
        print(monthly)
        prediction1 = pilreg.predict(poly.fit_transform([[monthly]]))
        previous = pilreg.predict(poly.fit_transform([[x[estimate]]]))
        #actual = pilreg.predict(poly.fit_transform([[months]]))
        #print(prediction1)
        #print(previous)
        prediction = abs(prediction1 - previous)
        print(prediction)
        pred = float(prediction)
        predicted_days = ((30*CREDIT)/pred)

        if predicted_days >=30:
            print('YOUR CREDIT CAN LAST FOR',(predicted_days), 'days. |YOUR REMAINING CREDIT CAN LAST FOR THE MONTH|')
        else:
            print('YOUR CREDIT CAN LAST FOR',(predicted_days), 'days. |WARNING.....YOUR CREDIT CANNOT LAST FOR A MONTH, PLEASE RECHARGE|')
        plt.scatter(x, y, edgecolors='red')
        plt.plot(x, pilreg.predict(X))
        plt.xlabel('CUMMULATIVE DAYS')
        plt.ylabel('CUMMULATIVE CONSUMPTION')
        plt.title((location))
        plt.show()








